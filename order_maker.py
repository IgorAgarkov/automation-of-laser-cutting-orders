# -*- coding: utf8 -*-
import openpyxl
import os
from shutil import copy2
from datetime import datetime
import re

if not os.path.exists('DXF'):                                           # создаём папку DXF, если она не существует
    os.makedirs('DXF')
    print('Папка "DXF" создана')
else:
    print('Папка "DXF" существует')

if not os.path.exists('Чертежи гибка PDF'):                             # создаём папку "Чертежи гибка", если она не существует
    os.makedirs('Чертежи гибка PDF')
    print('Папка Чертежи гибка PDF создана')
else:
    print('Папка "Чертежи гибка PDF" существует')

if not os.path.exists('Чертежи сварка PDF'):                            # создаём папку "Чертежи сварка", если она не существует
    os.makedirs('Чертежи сварка PDF')
    print('Папка Чертежи сварка PDF создана')
else:
    print('Папка "Чертежи сварка PDF" существует')



wb_fin = openpyxl.load_workbook(r'D:\Рабочая\!_Automation\laser_specification_final_template.xlsx')  # файл-шаблон результатов для лазера
sheet_fin = wb_fin['total']

wb_supply = openpyxl.load_workbook(r'D:\Рабочая\!_Automation\order_supply_template.xlsx')            # файл-шаблонр результаов для снабжения
sheet_supply = wb_supply.active


def getOrder(link):
    wb = openpyxl.load_workbook(link)
    ws_odr = wb['odr']
    ws_base = wb['base']
    i = 2
    orders_lst = []                                                     # итоговый список
    while ws_odr.cell(row = i, column = 2).value != None:  
        priduct = ws_odr.cell(row = i, column = 1).value
        j = 1
        prod = []
        while ws_base.cell(row = j, column = 1).value != priduct:       # ищем индекс в на вкладке base
            j += 1
        prod.append(ws_base.cell(row = j, column = 3).value)            # добавили ссылку на заявку на лазер в список
        prod.append(ws_odr.cell(row = i, column = 2).value)             # добавили количество
        prod.append(ws_odr.cell(row = i, column = 3).value)             # добавили материал
        prod.append(ws_base.cell(row = j, column = 4).value)            # добавили ссылку на заявку снабжению в список
        prod.append(ws_base.cell(row = j, column = 2).value)            # добавили ссылку на DXF в список
        prod.append(ws_base.cell(row = j, column = 5).value)            # добавили ссылку на чертежи гибка PDF
        prod.append(ws_base.cell(row = j, column = 6).value)            # добавили ссылку на чертежи сварка PDF 
        
        orders_lst.append(prod)
        i += 1
    return orders_lst


total_lazer = [[] for _ in range(8)]                                    # итоговая объединённая матрица из файлов excel для лазера
total_supply = [[] for _ in range(5)]                                   # итоговая объединённая матрица из файлов excel для снабжения


# собираем массив данных для заказа на резку
def copyRange(lst, total_tab):
    wb = openpyxl.load_workbook(lst[0])                                                            # файл-источник
    n = lst[1]                                                                                     # количество изделий
    ws = wb[lst[2]]                                                                                # вкладка в excel ("цинк", или "чёрный")
    i = 6                                                                                          # начиная со строки 6
    while ws.cell(row = i, column = 2).value != None:                                              # пока ячейки в колонке 2 не пустые выполняем:
        if ws.cell(row = i, column = 2).value not in total_tab[1]:                                 # если значения ячейки столбца 2 нет в списке выполняем:
            total_tab[1].append(ws.cell(row = i, column = 1 + 1).value)                            # заполняем значения колонки Наименование
            total_tab[2].append('=' + str(n) + '*' + str(ws.cell(row = i, column = 2 + 1).value))  # заполняем колонку "Количество" умножая значения на n
            for j in range(3, 8 ):                                                                 # оставшиеся колонки с 4 по 8 заполняем циклом
                total_tab[j].append(ws.cell(row = i, column = j + 1).value)
        else:
            ind = total_tab[1].index(ws.cell(row = i, column = 2).value)                                              # находим индекс уже имеющейся позиции
            total_tab[2][ind] = total_tab[2][ind] + '+' + str(n) + '*' + str(ws.cell(row = i, column = 2 + 1).value)  # прибавляем к содержимому ячейки количество повторяющейся позиции
        i += 1
    
    total_tab[0].clear()                                                # очищаем колонку с нумерацией
    for num in range(1, len(total_tab[1]) + 1):                         # проставляем нумерацию 
        total_tab[0].append(num)
    return total_tab


# записываем в файл excel массив данных на резку 
def writeRange(tab, sheetReceiving):
    for col in range(8):
        for row in range(len(tab[1])):
            sheetReceiving.cell(row = row + 6, column = col + 1).value = tab[col][row]

def writeSupply(tab, sheetReceiving):
    for col in range(5):
        for row in range(len(tab[1])):
            sheetReceiving.cell(row = row + 3, column = col + 1).value = tab[col][row]


# собираем массив данных для снабжения
def copySupply(lst, total_tab):
    if lst[3] != None:
        wb = openpyxl.load_workbook(lst[3])                                                            # файл-источник
        n = lst[1]                                                                                     # количество изделий
        ws = wb.active
        i = 3                                                                                          # начиная со строки 3
        while ws.cell(row = i, column = 2).value != None:                                              # пока ячейки в колонке 2 не пустые выполняем:
            if ws.cell(row = i, column = 2).value not in total_tab[1]:                                 # если значения ячейки столбца 2 нет в списке выполняем:
                total_tab[1].append(ws.cell(row = i, column = 1 + 1).value)                            # заполняем значения колонки Наименование
                total_tab[2].append('=' + str(n) + '*' + str(ws.cell(row = i, column = 2 + 1).value))  # заполняем колонку "Количество" умножая значения на n
                for j in range(3, 5 ):                                                                 # оставщиеся колонки с 4 по 5 заполняем циклом
                    total_tab[j].append(ws.cell(row = i, column = j + 1).value)
            else:
                ind = total_tab[1].index(ws.cell(row = i, column = 2).value)                           # находим индекс уже имеющейся позиции
                total_tab[2][ind] = total_tab[2][ind] + '+' + str(n) + '*' + str(ws.cell(row = i, column = 2 + 1).value)  # прибавляем к содержимому ячейки количество повторяющейся позиции
            i += 1
    
        total_tab[0].clear()                                            # очищаем колонку с нумерацией
        for num in range(1, len(total_tab[1]) + 1):                     # проставляем нумерацию 
            total_tab[0].append(num)
    return total_tab


def copyDXF(lst, goal_patch):                                           # для копирования файлов DXF
    src_patch = lst[4]
    print(src_patch)
    lst1 = os.listdir(src_patch)
    for dxf in lst1:
        goal_list = os.listdir(goal_patch)
        if dxf[-4:].lower() == '.dxf' and dxf not in goal_list:
            copy2(src_patch + '/' + dxf, goal_patch + '/' + dxf)    



def copyFiles(lst, ext, goal_patch):                                    # для копирования чертежей гибка PDF
    src_patch = lst[5]
    print(src_patch)
    lst1 = os.listdir(src_patch)
    for file in lst1:
        goal_list = os.listdir(goal_patch)
        if file[-4:].lower() == '.' + ext and file not in goal_list:
            copy2(src_patch + '/' + file, goal_patch + '/' + file) 

def copyFiles2(lst, ext, goal_patch):                                   # для копирования чертежей сварка PDF
    src_patch = lst[6]
    print(src_patch)
    lst1 = os.listdir(src_patch)
    for file in lst1:
        goal_list = os.listdir(goal_patch)
        if file[-4:].lower() == '.' + ext and file not in goal_list:
            copy2(src_patch + '/' + file, goal_patch + '/' + file) 




orders = getOrder('product_list.xlsx')                                  # вложенный список с информацией для обработки

for order in orders:                                                    # заполняем total_lazer через цикл по всем файлам excel
    copyRange(order, total_lazer)

for supply in orders:                                                   # заполняем total_supply через цикл по всем файлам excel
    copySupply(supply, total_supply)


directory_path = os.getcwd()
folder_name = os.path.basename(directory_path)                          # получаем имя дериктории
folder_name_lst = folder_name.split()                                   # список из имени директории

for s in folder_name_lst:                                               # получаем номер заказа из названия директории
    if s[0] == '№':
        order_num = s
        break
match = re.search(r'\d{4}.\d{2}.\d{2}', folder_name)                    # вычленяем дату из имени дериктории
date = datetime.strptime(match.group(), '%Y.%m.%d').date()
year = date.strftime("%Y")
month = date.strftime("%m")
day = date.strftime("%d")



file_name = f"Заявка на лазер {order_num} {date.strftime('%Y.%m.%d')}"


writeRange(total_lazer, sheet_fin)

sheet_fin.cell(row = 2, column = 3).value = order_num                   # вписываем номе рзаказа в финальный Excel на лазер
sheet_fin.cell(row = 3, column = 2).value = date.strftime('%d.%m.%Y')   # вписываем дату в финальный Excel на лазер

wb_fin.save(file_name + ".xlsx")


writeSupply(total_supply, sheet_supply)
wb_supply.save("snab.xlsx")

print()
print('Копируем DXF-файлы...')
for dx in orders:                                                       # копируем DXF-файлы в итоговую папку 'DXF' через цикл по всем ссылкам на dxf
    copyDXF(dx, 'DXF')
    
print()
print('Копируем "Чертежи гибка PDF"')
for pd in orders:                                                       # копируем PDF-файлы в итоговую папку 'Чертежи гибка PDF' через цикл по всем ссылкам на pdf
    copyFiles(pd, 'pdf', 'Чертежи гибка PDF')

print()
print('Копируем "Чертежи сварка PDF"')
for pd in orders:                                                       # копируем PDF-файлы в итоговую папку 'Чертежи сварка PDF' через цикл по всем ссылкам на pdf
    copyFiles2(pd, 'pdf', 'Чертежи сварка PDF')

print()
print('Выполнено!')
print()
input('Нажмите Enter, чтобы выйти')